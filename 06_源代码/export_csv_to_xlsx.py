import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def export_csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "识别结果"

    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file)
        for row in reader:
            ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column in ws.columns:
        max_length = 10
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 60))
        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(xlsx_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="将识别 CSV 结果导出为 Excel 文件")
    parser.add_argument("csv", help="输入 CSV 文件")
    parser.add_argument("xlsx", help="输出 XLSX 文件")
    args = parser.parse_args()

    export_csv_to_xlsx(Path(args.csv), Path(args.xlsx))
    print(f"Excel 已导出：{args.xlsx}")


if __name__ == "__main__":
    main()
